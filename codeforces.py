import time
from datetime import datetime
from urllib.error import HTTPError

import codeforces_api
import pytz
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from requests import RequestException


# 获取中国时区的时间
def get_cn_time(timestamp):
    dt_object_utc = datetime.utcfromtimestamp(timestamp).replace(tzinfo=pytz.utc)
    # 转换为中国时区
    cst = pytz.timezone('Asia/Shanghai')
    dt_object_cst = dt_object_utc.astimezone(cst)
    # print(str(dt_object_cst).split(' ')[0])
    dt_object_cst = dt_object_cst.replace(hour=0, minute=0, second=0)
    return dt_object_cst


name_col = 1
id_col = 2
cf_id_col = 3
cf_rating_col = 4

key = 'cbb429ffb396305d70f37818d9bc4f921cf61fc2'
secret = 'aaea4de623e5f9577f4448cdcde4d384a08e4edf'
# Authorized access to api.
cf_api = codeforces_api.CodeforcesApi(key, secret)

# 获取Codeforces所有比赛
all_contest = cf_api.contest_list()

# 导入队员基本信息表
basic_file = "队员信息表.xlsx"
basic_workbook = load_workbook(basic_file)

# 输入查询区间
start_str = input("请输入起始日期：\n")
pre = start_str.split('-')
start_date = get_cn_time(datetime(int(pre[0]), int(pre[1]), int(pre[2])).timestamp())
end_str = input("请输入结束日期：\n")
pre = end_str.split('-')
end_date = get_cn_time(datetime(int(pre[0]), int(pre[1]), int(pre[2])).timestamp())

# 选中信息工作表
total_sheet = basic_workbook.active

# 将要获取的所有比赛信息工作表（生成表）
competition_workbook = Workbook()
major_sheet = competition_workbook.active
major_sheet.title = 'ALL'
competition_workbook.save('CodeForces积分表 From' + start_str + 'to' + end_str + '.xlsx')

# 初始化生成表的基本信息并获取所有队员的CodeForcesID
major_sheet.cell(row=2, column=name_col, value='姓名')
major_sheet.cell(row=2, column=id_col, value='学号')
major_sheet.cell(row=2, column=cf_id_col, value='CodeforcesID')
major_sheet.cell(row=2, column=cf_rating_col, value='CodeforcesRating')

cf_id_of_member = []
for cell in total_sheet[1]:
    if cell.value == '姓名':
        for row in total_sheet.iter_rows(min_row=2, max_row=total_sheet.max_row, values_only=False):
            name_cell = row[cell.column - 1]
            if name_cell.value is not None:
                major_sheet.cell(row=name_cell.row + 1, column=name_col, value=name_cell.value)
    elif cell.value == '学号':
        for row in total_sheet.iter_rows(min_row=2, max_row=total_sheet.max_row, values_only=False):
            num_cell = row[cell.column - 1]
            if num_cell.value is not None:
                major_sheet.cell(row=num_cell.row + 1, column=id_col, value=num_cell.value)
    elif cell.value == 'CF_ID':
        for row in total_sheet.iter_rows(min_row=2, max_row=total_sheet.max_row, values_only=False):
            id_cell = row[cell.column - 1]
            if id_cell.value is not None:
                cf_id_of_member.append(str(id_cell.value))
                major_sheet.cell(row=id_cell.row + 1, column=cf_id_col, value=id_cell.value)
            else:
                cf_id_of_member.append('用户未填写')
                major_sheet.cell(row=id_cell.row + 1, column=cf_id_col, value='用户未填写')
competition_workbook.save('CodeForces积分表 From' + start_str + 'to' + end_str + '.xlsx')


# 枚举每一场具体比赛情况
for cur_contest in all_contest:
    col = major_sheet.max_column + 1
    cur_date = get_cn_time(cur_contest.start_time_seconds)

    if start_date <= cur_date <= end_date:
        competition_sheet = competition_workbook.create_sheet(title=cur_contest.name)
        competition_sheet.cell(row=2, column=name_col, value='姓名')
        competition_sheet.cell(row=2, column=id_col, value='学号')
        competition_sheet.cell(row=2, column=3, value='本场分数')

        # 获取本场比赛的题目集
        problem_set = cf_api.contest_standings(contest_id=cur_contest.id)['problems']
        # print(problem_set[0])

        # To Debug
        # standing = cf_api.contest_standings(str(cur_contest.id), handles=['Whxxxxx318'])['rows']
        # print(standing[0].problem_results[0].points)
        # for result in standing[0].problem_results:
        #     print(result.points)
        # print("=====")

        # 初始化每一场具体表格
        problem_col = 4
        for problem in problem_set:
            competition_sheet.cell(row=1, column=problem_col, value=problem.index)
            competition_sheet.cell(row=1, column=problem_col, value=problem.rating)
            problem_col += 1
            # print(problem.index)
            # print(problem.points)

        for cell in total_sheet[1]:
            if cell.value == '姓名':
                for row in total_sheet.iter_rows(min_row=2, max_row=total_sheet.max_row, values_only=False):
                    name_cell = row[cell.column - 1]
                    if name_cell.value is not None:
                        competition_sheet.cell(row=name_cell.row + 1, column=name_col, value=name_cell.value)
            elif cell.value == '学号':
                for row in total_sheet.iter_rows(min_row=2, max_row=total_sheet.max_row, values_only=False):
                    num_cell = row[cell.column - 1]
                    if num_cell.value is not None:
                        competition_sheet.cell(row=num_cell.row + 1, column=id_col, value=num_cell.value)

        # print(str(cur_contest.id) + " " + cur_contest.name)
        major_sheet.cell(row=2, column=col, value=cur_contest.name)

        # 对于每位队员查询这场比赛的具体情况
        cnt = 3
        for member in cf_id_of_member:
            try:
                standing = cf_api.contest_standings(str(cur_contest.id), handles=[member])['rows']
                # print(standing)
                if not standing:
                    major_sheet.cell(row=cnt, column=col, value='未参赛')
                    # print("No")
                    cnt += 1
                else:
                    # major_sheet.cell(row=cnt, column=col, value=standing[0].points)  # 比赛原本的题目分数
                    # print(standing[0].points)
                    ans = 0.
                    problem_col = 4
                    count = 0
                    for result in problem_set:
                        print(result.rating)
                        competition_sheet.cell(row=cnt, column=problem_col,
                                               value=(0 if standing[0].problem_results[count].points == 0 else 1))
                        ans += (((result.rating - 275) * (0 if standing[0].problem_results[count].points == 0 else 1)) / 300) ** 2
                        count += 1
                        problem_col += 1
                    competition_sheet.cell(row=cnt, column=3, value=ans)
                    major_sheet.cell(row=cnt, column=col, value=ans)
                    cnt += 1
            except Exception as e:
                major_sheet.cell(row=cnt, column=col, value='未参赛')
                cnt += 1
                pass

# 查找CodeForces Rating
cur_row = 3
col = major_sheet.max_column + 1
major_sheet.cell(row=2, column=col, value='区间内最高两场Rating加和')

# 查询区间内Rating加和
# 注意这里的rating_update_time_seconds与比赛开始时间并不一致，所以我们需要对查询区间做加一或减一的特殊处理
for member in cf_id_of_member:
    print([member])
    try:
        # contests_info = cf_api.contest_standings(contest_id='2026', handles=[member])
        v = cf_api.user_info([member])[0].rating
        if v is None:
            major_sheet.cell(row=cur_row, column=cf_rating_col, value='该用户未参与过比赛')
        else:
            major_sheet.cell(row=cur_row, column=cf_rating_col, value=v)
    except Exception as e:
        major_sheet.cell(row=cur_row, column=cf_rating_col, value='未查找到该用户名')
        cur_row += 1
        continue

    try:
        v = cf_api.user_rating(member)

        if v is not None:
            mx1 = 0
            mx2 = 0
            print(len(v))
            for item in v:
                # check_date = get_cn_time(cf_api.contest_standings(item.contest_id)['contest'].start_time_seconds)
                cur_date = get_cn_time(item.rating_update_time_seconds)
                print(cur_date)

                if start_date <= cur_date <= end_date:
                    print(item.new_rating)
                    if item.new_rating > mx1:
                        mx2 = mx1
                        mx1 = item.new_rating
                    elif item.new_rating > mx2:
                        mx2 = item.new_rating
            major_sheet.cell(row=cur_row, column=col, value=mx1 + mx2)
        cur_row += 1
    except Exception as e:
        cur_row += 1
        continue

    # try:
    #     url = 'https://codeforces.com/contests/with/' + member
    #     data = pd.read_html(url)
    #     print(data[0])
    # except Exception as e:
    #     cur_row += 1
    #     continue
    # print(contests_info['rows'])

# 计算分数总和 与 区间内两场最高加和
right_col = major_sheet.max_column + 1
major_sheet.cell(row=2, column=right_col, value='分数总和')
major_sheet.cell(row=2, column=right_col + 1, value='区间内最高两场分数加和')
max_row = 3
for row in major_sheet.iter_rows(min_row=3, max_row=major_sheet.max_row, values_only=True):
    ans = 0.
    mx1 = 0
    mx2 = 0
    for col in range(cf_rating_col + 1, right_col - 1):
        val = major_sheet.cell(max_row, col).value
        print(val)
        if val is not None and val != '未参赛':
            # print(str(max_row) + " " + str(col))
            ans += val
            if val > mx1:
                mx2 = mx1
                mx1 = val
            elif val > mx2:
                mx2 = val

    major_sheet.cell(row=max_row, column=right_col, value=ans)
    major_sheet.cell(row=max_row, column=right_col + 1, value=mx1 + mx2)

    max_row += 1

competition_workbook.save('CodeForces积分表 From' + start_str + 'to' + end_str + '.xlsx')
