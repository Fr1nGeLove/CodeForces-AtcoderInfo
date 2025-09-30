import re
import time
from urllib.error import HTTPError

import requests
import pandas as pd
from datetime import datetime
from openpyxl.workbook import Workbook
from requests import RequestException
from openpyxl import load_workbook

# 一般只需要改 contest_name 即可
# 如果 Cookie 失效需要更换 Cookie
# 需要开启 VPN
# 如果 VPN 监听端口不同，需要修改 proxies
# contest_name = "abc359"

# Request Headers Cookie
headers = {
    'Cookie': '_ga=GA1.1.1021690899.1718458161; language=en; __pp_uid=fCDJaH8JyX66TjHIjF4Nxvtvpf4fxOlu; REVEL_FLASH=; REVEL_SESSION=d18a8aa4cd6b31757995619e55f0cd43212741c3-%00SessionKey%3A405d44fb303c68ea84c9dbe309a10a28a57cc5e7b1e645716e029d1259400d84%00%00UserScreenName%3AFr1nGeLove%00%00UserName%3AFr1nGeLove%00%00a%3Afalse%00%00w%3Afalse%00%00csrf_token%3AOh5dSHML0U%2BvAjU2SQBfO5u1t7S75EhsekBR5L0Flw8%3D%00%00_TS%3A1754650654%00; _ga_RC512FD18N=GS1.1.1739098609.26.1.1739098646.0.0.0'
}
proxies = {
    'http': 'http://127.0.0.1:7890',
    'https': 'http://127.0.0.1:7890',
}

atc_id_column = None
name_col = 1
id_col = 2
atc_id_col = 3
atc_rating_col = 4

# 设置重试次数和超时时间
max_retries = 3
timeout = 5  # 超时时间，单位秒
retries = 0

# 导入队员基本信息表
basic_file = "test.xlsx"
basic_workbook = load_workbook(basic_file)

# 选择工作表
total_sheet = basic_workbook.active

competition = []

# query_type = input("请输入查询类型：（A）时间段（B）场次\n")

# if query_type == 'B':
#     in_str = input("请输入查询场次，若要查询多场请用空格隔开。\n")
#     competition = in_str.split(' ')
# else:

# 输入查询区间
start_str = input("请输入起始日期：\n")
pre = start_str.split('-')
start_date = datetime(int(pre[0]), int(pre[1]), int(pre[2]))
end_str = input("请输入结束日期：\n")
pre = end_str.split('-')
end_date = datetime(int(pre[0]), int(pre[1]), int(pre[2]))

# 枚举每一页比赛
for page_num in range(1, 7):
    url = 'https://atcoder.jp/contests/archive?page=' + str(page_num)
    data = pd.read_html(url)
    row_cnt = 0
    # 找出要查询的所有场次
    for date in data[0]['Start Time (local time)']:
        pre = date.split(' ')[0].split('-')
        check_date = datetime(int(pre[0]), int(pre[1]), int(pre[2]))
        competition_name = data[0]['Contest Name'][row_cnt]
        # print(competition_name)

        if start_date <= check_date <= end_date:
            if competition_name.find("AtCoder Beginner Contest") != -1:
                # print(competition_name.find("AtCoder Beginner Contest"))
                div = competition_name.split(' ')
                # print(div)
                # print(re.findall(r'\d+', competition_name))
                # num = str(re.findall(r'\d+', competition_name)[-1])
                num = str(re.findall(r'\d+', str(div[div.index('Beginner') + 2]))[-1])
                competition.append("abc" + num)
                print("abc" + num)
            elif competition_name.find("AtCoder Grand Contest") != -1:
                # num = str(re.findall(r'\d+', competition_name)[-1])
                div = competition_name.split(' ')
                # print(div)
                # print(re.findall(r'\d+', competition_name))
                # num = str(re.findall(r'\d+', competition_name)[-1])
                num = str(re.findall(r'\d+', str(div[div.index('Grand') + 2]))[-1])
                competition.append("agc" + num)
                print("agc" + num)
            elif competition_name.find("AtCoder Regular Contest") != -1:
                # num = str(re.findall(r'\d+', competition_name)[-1])
                div = competition_name.split(' ')
                # print(div)
                # print(re.findall(r'\d+', competition_name))
                # num = str(re.findall(r'\d+', competition_name)[-1])
                num = str(re.findall(r'\d+', str(div[div.index('Regular') + 2]))[-1])
                print("arc" + num)
                competition.append("arc" + num)
        row_cnt += 1

print(competition)
# 将要获取的所有比赛信息工作表（生成表）
competition_workbook = Workbook()
major_sheet = competition_workbook.active
major_sheet.title = 'ALL'
competition_workbook.save('Atcoder积分表 From' + start_str + 'to' + end_str + '.xlsx')

tag = False
# 依次查询每一场比赛
for contest_name in competition:
    url = "https://atcoder.jp/contests/" + contest_name + "/standings/json"
    score = {}
    # 设置重试次数和超时时间
    max_retries = 3
    timeout = 5  # 超时时间，单位秒
    retries = 0
    RankInfo = {}

    # 请求连接
    while retries < max_retries:
        try:
            # response = requests.get(url, headers=headers, timeout=timeout)
            response = requests.get(url, headers=headers, proxies=proxies, timeout=timeout)
            response.raise_for_status()
            RankInfo = response.json()
            break
        except RequestException as e:
            print(f"请求失败: {e}, 正在重试（剩余次数：{max_retries - retries - 1}）...")
            retries += 1
            time.sleep(1)

    if retries == max_retries:
        print("已达到最大重试次数，请求失败。")
    else:
        AcceptInfo = {}

        # 题目放置列号
        problem = {}
        index = 5
        for item in RankInfo["TaskInfo"]:
            problem[item["Assignment"]] = index
            score[item["Assignment"]] = 0
            index += 1

        # 获取当场比赛每道题目的具体分数
        for problem_key, problem_value in RankInfo["StandingsData"][0]["TaskResults"].items():
            problem_id = problem_key.split("_")[1].upper()
            score[problem_id] = problem_value["Score"] // 100

        print(problem)
        print(score)
        for item in RankInfo["StandingsData"]:
            AcceptInfo[item["UserScreenName"]] = {
                "Rating": item["Rating"],
                "AcceptNum": item["TotalResult"]["Accepted"],
                "AcceptProblem": {}
            }
            for problem_key, problem_value in item["TaskResults"].items():
                problem_id = problem_key.split("_")[1].upper()
                if problem_value["Status"] == 1:
                    AcceptInfo[item["UserScreenName"]]["AcceptProblem"][problem_id] = 1

        competition_sheet = competition_workbook.create_sheet(title=contest_name.upper())

        # 初始化每场比赛的队员信息
        competition_sheet.cell(row=2, column=4, value='本场分数')
        for problem_id, problem_column in problem.items():
            competition_sheet.cell(row=2, column=problem_column, value=problem_id)
            competition_sheet.cell(row=1, column=problem_column, value=score[problem_id])
        for cell in total_sheet[1]:
            if cell.value == '姓名':
                major_sheet.cell(row=2, column=name_col, value='姓名')
                competition_sheet.cell(row=2, column=2, value='姓名')
                for row in total_sheet.iter_rows(min_row=2, max_row=total_sheet.max_row, values_only=False):
                    name_cell = row[cell.column - 1]
                    if name_cell.value is not None:
                        competition_sheet.cell(row=name_cell.row + 1, column=2, value=name_cell.value)
                        if tag is False:
                            major_sheet.cell(row=name_cell.row + 1, column=name_col, value=name_cell.value)
            elif cell.value == '学号':
                competition_sheet.cell(row=2, column=3, value='学号')
                major_sheet.cell(row=2, column=id_col, value='学号')
                for row in total_sheet.iter_rows(min_row=2, max_row=total_sheet.max_row, values_only=False):
                    num_cell = row[cell.column - 1]
                    if num_cell.value is not None:
                        competition_sheet.cell(row=num_cell.row + 1, column=3, value=num_cell.value)
                        if tag is False:
                            major_sheet.cell(row=num_cell.row + 1, column=id_col, value=num_cell.value)
            elif cell.value == 'ATC_ID':
                major_sheet.cell(row=2, column=atc_id_col, value='AtcoderID')
                for row in total_sheet.iter_rows(min_row=2, max_row=total_sheet.max_row, values_only=False):
                    id_cell = row[cell.column - 1]
                    if tag is False:
                        if id_cell.value is not None:
                            major_sheet.cell(row=id_cell.row + 1, column=atc_id_col, value=id_cell.value)
                        else:
                            major_sheet.cell(row=id_cell.row + 1, column=atc_id_col, value='用户未填写')
                tag = True

        for cell in total_sheet[1]:  # 遍历第一行
            if cell.value == 'ATC_ID':
                atc_id_column = cell.column  # 获取列字母（如'A'、'B'等）
                break  # 找到后退出循环

        # atc_rating_column = None
        # for cell in total_sheet[1]:
        #     if cell.value == 'atc_rating':
        #         atc_rating_column = cell.column
        #         break

        name_column = None
        for cell in total_sheet[1]:
            if cell.value == '姓名':
                name_column = cell.column
                break

        name2_column = None
        for cell in competition_sheet[2]:
            if cell.value == '姓名':
                name2_column = cell.column
                break

        # 如果没有找到'atc_id'，则打印错误消息
        if atc_id_column is None or name_column is None:
            print("AtcoderID or AtcoderRating or Name not found.")
        else:
            # total_sheet.append([None] * total_sheet.max_column)
            # total_sheet.cell(row=1, column=total_sheet.max_column + 1, value=contest_name.upper())
            major_sheet.cell(row=2, column=major_sheet.max_column + 1, value=contest_name.upper())
            max_row = 1
            NotFind = []
            # 遍历该列的所有单元格（从第二行开始）
            for row in total_sheet.iter_rows(min_row=2, max_row=total_sheet.max_row, values_only=False):
                # 获取该列的单元格值
                atc_id_value = row[atc_id_column - 1].value  # 注意索引从0开始，所以要减1
                name_value = row[name_column - 1].value
                max_row += 1
                if atc_id_value is None:
                    continue
                atc_id_value = str(atc_id_value).replace(" ", "")
                info_item = AcceptInfo.get(atc_id_value, None)
                print(info_item)
                if info_item is None:
                    NotFind.append(name_value)
                    # total_sheet.cell(row=max_row, column=total_sheet.max_column, value=None)
                else:
                    AcceptNum = info_item.get('AcceptNum', None)
                    Rating = info_item.get('Rating', None)
                    print("Rating:", Rating, "|| num:", AcceptNum, "|| AtcoderID:", atc_id_value)
                    # total_sheet.cell(row=max_row, column=total_sheet.max_column, value=AcceptNum)
                    # total_sheet.cell(row=max_row, column=atc_rating_column, value=Rating)  # 更新Atcoder的Rating
                    row_id = 1

                    # 标记通过情况
                    for row2 in competition_sheet.iter_rows(min_row=2, max_row=competition_sheet.max_row,
                                                            values_only=False):
                        ans = 0.
                        row_id += 1
                        if row2[name2_column - 1].value == row[name_column - 1].value:
                            for problem_id, problem_column in problem.items():
                                if info_item["AcceptProblem"].get(problem_id, None) is None:
                                    competition_sheet.cell(row=row_id, column=problem_column, value=0)
                                else:
                                    competition_sheet.cell(row=row_id, column=problem_column, value=1)
                                    ans += (score[problem_id] / 100 - 1) ** 2  # 分数计算

                            competition_sheet.cell(row=row_id, column=4, value=ans)
                            major_sheet.cell(row=row_id, column=major_sheet.max_column, value=ans)

            # basic_workbook.save("2024暑假集训选拔new.xlsx")
            print("save success")
            print("未参赛名单：")
            print(NotFind)

    competition_workbook.save('Atcoder积分表 From' + start_str + 'to' + end_str + '.xlsx')
# sheet_to_remove = competition_workbook['Sheet']
# competition_workbook.remove(sheet_to_remove)
right_col = major_sheet.max_column + 1
# print(right_col)

# 计算分数总和 与 区间内两场最高加和
major_sheet.cell(row=2, column=right_col, value='分数总和')
major_sheet.cell(row=2, column=right_col + 1, value='区间内最高两场分数加和')
max_row = 3
for row in major_sheet.iter_rows(min_row=3, max_row=major_sheet.max_row, values_only=True):
    ans = 0.
    mx1 = 0
    mx2 = 0
    for col in range(atc_rating_col + 1, right_col):
        val = major_sheet.cell(max_row, col).value
        if val is not None:
            # print(str(max_row) + " " + str(col))
            ans += val
            if val > mx1:
                mx2 = mx1
                mx1 = val
            elif val > mx2:
                mx2 = val
        else:
            major_sheet.cell(row=max_row, column=col, value='未参赛')

    major_sheet.cell(row=max_row, column=right_col, value=ans)
    major_sheet.cell(row=max_row, column=right_col + 1, value=mx1 + mx2)

    max_row += 1

competition_workbook.save('Atcoder积分表 From' + start_str + 'to' + end_str + '.xlsx')

# 查Rating
# major_sheet.cell(row=2, col=1, value='AtcoderRating')
# if query_type == 'A':
major_sheet.cell(row=2, column=major_sheet.max_column + 1, value='区间内最高两场Rating加和')
major_sheet.cell(row=2, column=atc_rating_col, value='AtcoderRating')

max_row = 3
for row in total_sheet.iter_rows(min_row=2, max_row=total_sheet.max_row, values_only=False):
    atc_id_value = row[atc_id_column - 1].value  # 注意索引从0开始，所以要减1
    if atc_id_value is None:
        # major_sheet.cell(row=max_row, column=3, value='该用户未填写AtcoderID')
        major_sheet.cell(row=max_row, column=atc_rating_col, value='该用户未填写AtcoderID')
        max_row += 1
        continue
    # major_sheet.cell(row=max_row, column=atc_id_col, value=atc_id_value)
    url = 'https://atcoder.jp/users/' + atc_id_value + '/history'
    print(url)
    try:
        data = pd.read_html(url)
    except HTTPError as e:
        major_sheet.cell(row=max_row, column=atc_rating_col, value='未查找到该用户')
        major_sheet.cell(row=max_row, column=major_sheet.max_column, value='未查找到该用户')
        max_row += 1
        continue
    except Exception as e:
        major_sheet.cell(row=max_row, column=atc_rating_col, value='该用户未进行任何一场比赛')
        major_sheet.cell(row=max_row, column=major_sheet.max_column, value='该用户未进行任何一场比赛')
        max_row += 1
        continue

    print(data[0]['New Rating'])
    row_cnt = 0
    mx1 = 0
    mx2 = 0
    rating = 0
    for date in data[0]['Date']:
        pre = date.split(' ')[0].split('-')
        check_date = datetime(int(pre[0]), int(pre[1]), int(pre[2]))
        competition_name = data[0]['Contest'][row_cnt]
        if data[0]['New Rating'][row_cnt] == '-':
            row_cnt += 1
            continue
        this_rating = int(data[0]['New Rating'][row_cnt])
        row_cnt += 1
        if start_date <= check_date <= end_date:
            if mx1 < this_rating:
                mx2 = mx1
                mx1 = this_rating
            elif mx2 < this_rating:
                mx2 = this_rating
        rating = this_rating
    major_sheet.cell(row=max_row, column=major_sheet.max_column, value=mx1 + mx2)
    major_sheet.cell(row=max_row, column=atc_rating_col, value=rating)

    max_row += 1
    competition_workbook.save('Atcoder积分表 From' + start_str + 'to' + end_str + '.xlsx')
competition_workbook.save('Atcoder积分表 From' + start_str + 'to' + end_str + '.xlsx')
